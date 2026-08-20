"""Relatório de homologação do RPA1 — quais e-mails foram lidos e o que cada
um resultou.

Pedido da homologação de 2026-08-18: uma planilha com assunto/remetente/data
de cada e-mail e o que aconteceu com ele — capturado e identificado como
Detraf/despesa (com qual operadora), capturado mas reprovado na validação
(com o motivo), operadora não identificada, ou nem capturado (ficou em
"Detraf Despesas" porque não passou no filtro de negócio, ou já tinha sido
rastreado antes e foi pulado).

**Só para homologação** — não faz parte do fluxo do robô, não é chamado por
nenhum `main.py`. Roda depois de uma execução do RPA1, na mesma máquina
(precisa do Outlook Desktop Classic aberto, logado na conta OUTLOOK_ACCOUNT) e
do mesmo `.env` que o RPA1 usou, para achar `_rastreamento.json` e as pastas de
destino (Operadoras, Quarentena, Não Identificados).

Uso::

    python gerar_relatorio_emails_rpa1.py
    python gerar_relatorio_emails_rpa1.py --saida relatorio_202607.xlsx

Uma linha por e-mail; se o e-mail trouxe mais de um arquivo, uma linha por
arquivo (assunto/remetente/data repetidos).
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

_RAIZ = Path(__file__).resolve().parent
for _caminho in (_RAIZ, _RAIZ / "rpa1_captura"):
    if str(_caminho) not in sys.path:
        sys.path.insert(0, str(_caminho))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from comum.config import configuration
from comum.integracoes.outlook import OutlookService
from comum.integracoes.outlook_config import OutlookConfig


@dataclass
class LinhaRelatorio:
    assunto: str
    remetente: str
    recebido_em: str
    situacao_email: str
    arquivo: str = ""
    status_arquivo: str = ""
    detalhe: str = ""


def _para_data(valor) -> date | None:
    """Extrai só a data (sem hora/fuso) de um `ReceivedTime` do COM."""
    if not valor:
        return None
    try:
        return date(valor.year, valor.month, valor.day)
    except AttributeError:
        return None


def _formatar_data(valor) -> str:
    if not valor:
        return ""
    if isinstance(valor, str):
        try:
            valor = datetime.fromisoformat(valor)
        except ValueError:
            return valor
    try:
        return valor.strftime("%d/%m/%Y %H:%M")
    except AttributeError:
        return str(valor)


def _carregar_rastreamento() -> dict[str, list[dict]]:
    """`{entry_id: [registros de arquivo baixado daquele e-mail]}`."""
    caminho = Path(configuration.RASTREAMENTO_ARQUIVO_PATH)
    if not caminho.is_file():
        print(f"[aviso] rastreamento não encontrado em {caminho} — "
              "e-mails capturados vão aparecer sem o(s) arquivo(s).")
        return {}

    try:
        registros = json.loads(caminho.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as erro:
        print(f"[aviso] falha ao ler rastreamento: {erro}")
        return {}

    por_entry_id: dict[str, list[dict]] = {}
    for registro in registros:
        por_entry_id.setdefault(registro["entry_id"], []).append(registro)
    return por_entry_id


#: (rótulo do status, raiz onde procurar, é a árvore de operadoras?)
_RAIZES_DE_BUSCA = (
    ("Aprovado", configuration.CAMINHO_OPERADORAS, True),
    ("Reprovado na validação", configuration.DIRETORIO_QUARENTENA, False),
    ("Operadora não identificada", configuration.DIRETORIO_NAO_IDENTIFICADOS, False),
)


def _localizar_arquivo(nome_arquivo: str) -> tuple[str, str]:
    """
    Procura `nome_arquivo` nas pastas de destino do RPA1.

    Returns:
        `(status, detalhe)` — detalhe é a operadora (se aprovado) ou o motivo
        da recusa lido do `*_RECUSADO.md` ao lado (se reprovado).
    """
    for status, raiz, eh_arvore_operadoras in _RAIZES_DE_BUSCA:
        if not raiz:
            continue
        raiz = Path(raiz)
        if not raiz.is_dir():
            continue

        try:
            encontrados = list(raiz.rglob(nome_arquivo))
        except OSError:
            continue

        if not encontrados:
            continue

        encontrado = encontrados[0]
        detalhe = ""
        if eh_arvore_operadoras:
            try:
                detalhe = encontrado.relative_to(raiz).parts[0]
            except (ValueError, IndexError):
                detalhe = ""
        else:
            md = encontrado.with_name(f"{encontrado.stem}_RECUSADO.md")
            if md.is_file():
                try:
                    detalhe = md.read_text(encoding="utf-8").strip()
                except OSError:
                    detalhe = ""
        return status, detalhe

    return "Não encontrado (ainda não processado nesta máquina, ou já removido)", ""


def _buscar_subpasta_sem_criar(parent_folder, nome: str, tentativas: int = 3):
    """
    Procura uma subpasta pelo nome (case-insensitive) — **nunca cria**.

    Script de leitura não deve ter efeito colateral de criar pasta no
    Outlook; isso é papel do robô, não de um relatório de conferência. A
    listagem de `Folders` às vezes não mostra uma subpasta que existe de
    verdade (visto antes com "PROCESSADOS" — some sozinho depois de um
    tempo, provável atraso de sincronização do cache do Exchange) — por
    isso tenta de novo algumas vezes antes de desistir, em vez de assumir
    que "não encontrada na primeira vez" significa "não existe".

    Returns:
        A subpasta, ou `None` se não achou depois de todas as tentativas.
    """
    for tentativa in range(1, tentativas + 1):
        for i in range(1, parent_folder.Folders.Count + 1):
            pasta = parent_folder.Folders.Item(i)
            if pasta.Name.lower() == nome.lower():
                return pasta
        if tentativa < tentativas:
            print(f"[aviso] pasta '{nome}' não apareceu na tentativa {tentativa} — tentando de novo...")
            time.sleep(2)
    return None


def montar_linhas(
    desde: date | None = None, ate: date | None = None
) -> list[LinhaRelatorio]:
    """
    Args:
        desde: Só inclui e-mails recebidos a partir desta data (inclusive).
        ate: Só inclui e-mails recebidos até esta data (inclusive).
    """
    config = OutlookConfig.from_configuration()
    outlook = OutlookService(config.account)
    rastreamento = _carregar_rastreamento()

    pasta_detraf = outlook._get_or_create_top_level_folder(  # noqa: SLF001 — script de diagnóstico
        config.detraf_despesas_folder
    )
    pasta_processados = _buscar_subpasta_sem_criar(pasta_detraf, config.processados_folder)
    if pasta_processados is None:
        print(
            f"[aviso] pasta '{config.processados_folder}' não encontrada dentro de "
            f"'{config.detraf_despesas_folder}' — relatório vai sair sem os e-mails já "
            "capturados. Rode de novo em alguns minutos."
        )

    linhas: list[LinhaRelatorio] = []

    def _processar_pasta(pasta, situacao_quando_sem_rastreamento: str) -> None:
        total = pasta.Items.Count
        for i in range(1, total + 1):
            item = pasta.Items.Item(i)
            try:
                if item.Class != 43:  # olMailItem
                    continue
                entry_id = item.EntryID
                assunto = item.Subject or ""
                remetente = item.SenderEmailAddress or ""
                data_recebimento = _para_data(item.ReceivedTime)
                recebido_em = _formatar_data(item.ReceivedTime)
            except Exception as erro:  # noqa: BLE001 — item exótico, não trava o relatório
                linhas.append(LinhaRelatorio(
                    assunto="(falha ao ler este item)",
                    remetente="",
                    recebido_em="",
                    situacao_email=f"Erro ao ler: {erro}",
                ))
                continue

            if desde and data_recebimento and data_recebimento < desde:
                continue
            if ate and data_recebimento and data_recebimento > ate:
                continue

            arquivos = rastreamento.get(entry_id, [])

            if not arquivos:
                linhas.append(LinhaRelatorio(
                    assunto=assunto,
                    remetente=remetente,
                    recebido_em=recebido_em,
                    situacao_email=situacao_quando_sem_rastreamento,
                ))
                continue

            for registro in arquivos:
                nome_arquivo = Path(registro["caminho_arquivo"]).name
                status, detalhe = _localizar_arquivo(nome_arquivo)
                linhas.append(LinhaRelatorio(
                    assunto=assunto,
                    remetente=remetente,
                    recebido_em=recebido_em,
                    situacao_email="Capturado",
                    arquivo=nome_arquivo,
                    status_arquivo=status,
                    detalhe=detalhe,
                ))

    # E-mails já capturados (movidos pra PROCESSADOS em alguma execução).
    if pasta_processados is not None:
        _processar_pasta(pasta_processados, situacao_quando_sem_rastreamento=(
            "Em PROCESSADOS mas sem registro de rastreamento (verificar)"
        ))

    # E-mails que ainda estão em "Detraf Despesas" — nunca foram capturados
    # (não passaram no filtro de negócio), ou já estavam rastreados de uma
    # execução anterior e por isso foram pulados sem mover.
    _processar_pasta(pasta_detraf, situacao_quando_sem_rastreamento=(
        "Não capturado — não passou no filtro de negócio (sem anexo Detraf "
        "relevante), ou está aguardando a próxima execução"
    ))

    return linhas


def gerar_planilha(linhas: list[LinhaRelatorio], caminho_saida: Path) -> None:
    livro = Workbook()
    aba = livro.active
    aba.title = "E-mails RPA1"

    cabecalho = [
        "Assunto", "Remetente", "Recebido em", "Situação do e-mail",
        "Arquivo", "Status do arquivo", "Operadora / Motivo",
    ]
    aba.append(cabecalho)
    for celula in aba[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="4472C4")
        celula.alignment = Alignment(wrap_text=True)

    cores_status = {
        "Capturado": "C6EFCE",
        "Aprovado": "C6EFCE",
        "Reprovado na validação": "FFC7CE",
        "Operadora não identificada": "FFEB9C",
    }

    for linha in linhas:
        aba.append([
            linha.assunto, linha.remetente, linha.recebido_em,
            linha.situacao_email, linha.arquivo, linha.status_arquivo,
            linha.detalhe,
        ])
        cor = cores_status.get(linha.status_arquivo) or cores_status.get(linha.situacao_email)
        if cor:
            aba.cell(row=aba.max_row, column=6).fill = PatternFill("solid", fgColor=cor)

    larguras = [40, 30, 16, 40, 35, 26, 60]
    for indice, largura in enumerate(larguras, start=1):
        aba.column_dimensions[aba.cell(row=1, column=indice).column_letter].width = largura
    aba.freeze_panes = "A2"

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    livro.save(caminho_saida)


def _tipo_data(texto: str) -> date:
    try:
        return datetime.strptime(texto, "%Y-%m-%d").date()
    except ValueError as erro:
        raise argparse.ArgumentTypeError(
            f"'{texto}' não é uma data válida — use AAAA-MM-DD."
        ) from erro


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--desde", type=_tipo_data, default=None,
        help="Só inclui e-mails recebidos a partir desta data (AAAA-MM-DD).",
    )
    parser.add_argument(
        "--ate", type=_tipo_data, default=None,
        help="Só inclui e-mails recebidos até esta data (AAAA-MM-DD).",
    )
    parser.add_argument(
        "--saida", type=Path, default=None,
        help="Caminho do .xlsx de saída. Padrão: nome gerado com a data/período.",
    )
    args = parser.parse_args()

    saida = args.saida
    if saida is None:
        sufixo_periodo = ""
        if args.desde or args.ate:
            sufixo_periodo = f"_{args.desde or 'inicio'}_a_{args.ate or 'hoje'}"
        saida = Path(
            f"relatorio_emails_rpa1_{datetime.now():%Y%m%d_%H%M%S}{sufixo_periodo}.xlsx"
        )

    if args.desde and args.ate and args.desde > args.ate:
        parser.error("--desde não pode ser depois de --ate.")

    print("Conectando ao Outlook e lendo 'Detraf Despesas' + 'PROCESSADOS'...")
    if args.desde or args.ate:
        print(f"Filtrando por período: {args.desde or '(sem início)'} a {args.ate or '(sem fim)'}.")
    linhas = montar_linhas(desde=args.desde, ate=args.ate)
    print(f"{len(linhas)} linha(s) montada(s). Gravando em {saida}...")
    gerar_planilha(linhas, saida)
    print(f"Pronto: {saida.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
