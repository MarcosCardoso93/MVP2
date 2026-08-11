import os
import pandas as pd
import openpyxl
from datetime import datetime

from src.config.config import (
    DIRETORIO_RELATORIO_BAIXADO, DIRETORIO_ENCONTRO_CONTAS, DIRETORIO_INCONSISTENCIAS,
    CELULA_SUBTOTAL_DESPESA, PERIODO_REF, CONFIG_WEBFAT,
)
#from src.config.conexao import Banco
from src.services.AGI.AGI_config import AGI_CONFIG

"""
HU-20 - Verificacao do Relatorio Receitas e Despesas no AGI
=============================================================
REAPROVEITADO do exemplo RPA_DETRAF_RECEITA:
    - A navegacao "AGI > Relatorios > Detraf > Receitas e Despesas", o filtro por PERIODO
      e a exportacao para CSV sao EXATAMENTE o metodo AGI_CONFIG.Baixar_Remessa() (ja
      copiado sem alteracao para este pacote). Ele ja cobre o primeiro criterio de aceite:
      "Acesso a Relatorios > Detraf > Receitas e Despesas" + "Filtro por periodo".
    - A logica de comparacao (filtrar coluna, somar Vlr. Bruto por operadora, cruzar com o
      Encontro de Contas, calcular diferenca) foi adaptada de
      src/services/AGI/Criacao_Remessa.py -> metodo _validar_valores(). La ele filtra
      Natureza == 'C' (credito/receita); aqui filtramos Natureza == 'D' (debito/despesa),
      conforme HU-20 pede.

ADAPTADO (funciona diferente do exemplo):
    - Natureza 'C' -> 'D'.
    - CONFIRMADO (22/07, prints reais do As Is - imagens 3/12/8/14): o export de
      "Relatorios > Detraf > Receitas e Despesas" tem EXATAMENTE as mesmas colunas usadas
      no exemplo de Receita ("Grp. Oper .Prest", "Natureza", "Vlr. Bruto" etc.) - e o MESMO
      relatorio para os dois fluxos, so muda o filtro de Natureza (C vs D). Nao precisa mais
      de confirmacao - o mapeamento de coluna abaixo e o mesmo do exemplo.
    - CONFIRMADO: o Encontro de Contas E uma planilha Excel real, com uma ABA POR OPERADORA
      (ex.: aba "AMPERNET") e o "Subtotal Despesa" numa celula fixa (linha 87, confirmado
      no print - ver CELULA_SUBTOTAL_DESPESA no config.py). NAO e um CSV agregável por
      groupby como no exemplo de Receita - a leitura e por aba (nome da operadora) + celula
      fixa via openpyxl, implementada abaixo.
    - Alternativa observada no print (imagem 14 do As Is): a propria grid do AGI tem um
      recurso nativo de agregacao por coluna (dropdown "Sum" no cabecalho de "Vlr. Bruto",
      que mostra o total direto na tela). Pode servir como conferencia visual/manual, mas
      para o robo a soma via pandas (abaixo) e mais confiavel que tentar ler o valor
      agregado da tela por imagem/OCR.

NOVO (nao existe equivalente no exemplo):
    - Filtro por OPERADORA dentro do laco de verificacao (o exemplo comparava tudo de uma
      vez via groupby; aqui a HU pede "repete para todas as operadoras", o que o groupby
      ja cobre, mas falta decidir se cada operadora gera uma linha de log/alerta separada).
    - Mecanismo de "sinalizacao de inconsistencia": o exemplo so escrevia um Excel local
      (Controle_Valores_Receita_*.xlsx) sem nenhum alerta. Aqui e preciso decidir COMO
      sinalizar (Webfat, e-mail, log em tabela) - ver README.
"""


class Verificacao_Relatorio:

    def __init__(self):
        self.AGI_CONFIG = AGI_CONFIG()
        #self.db = Banco(CONFIG_WEBFAT)
        self.periodo = PERIODO_REF

    # ------------------------------------------------------------------
    # Fluxo principal
    # ------------------------------------------------------------------
    def Fluxo_Verificacao_Relatorio(self):
        # Estrutura de log ja pronta (mesmo padrao do RPA_DETRAF_RECEITA, so mudando a tabela
        # em conexao.py) - descomentar quando tbl_rpa_log_detraf_despesa_verificacao for
        # criada/confirmada. O try/except abaixo so existe pra dar um lugar pro log de erro;
        # o comportamento atual (deixar a excecao subir) continua identico via o "raise".
        # self.db.log(processo="Iniciado", status="Sucesso", descricao="Inicio da verificacao do relatorio")
        try:
            self.AGI_CONFIG.Fechar_AGI()
            self.AGI_CONFIG.Inicializando_AGI()
            self.AGI_CONFIG.Acessando_producao_AGI()
            self.AGI_CONFIG.Login_AGI_producao()

            # REAPROVEITADO 100%: navega Relatorios > Detraf > Receitas e Despesas, filtra por
            # periodo e exporta CSV (metodo pronto em AGI_config.py, ja usado no exemplo).
            self.AGI_CONFIG.Baixar_Remessa()

            self.AGI_CONFIG.Fechar_AGI()

            df_divergencias = self._comparar_com_encontro_de_contas()
            self._sinalizar_inconsistencias(df_divergencias)
            # self.db.log(processo="Finalizado", status="Sucesso", descricao="Verificacao concluida")
        except Exception as e:
            # self.db.log(processo="Finalizado", status="Erro", descricao=str(e))
            raise

    # ------------------------------------------------------------------
    # ADAPTADO de Criacao_Remessa._validar_valores() - trocado Natureza 'C' -> 'D'.
    # Nomes de coluna do export CONFIRMADOS (identicos ao exemplo de Receita - mesmo
    # relatorio do AGI serve para os dois fluxos).
    # ------------------------------------------------------------------
    def _comparar_com_encontro_de_contas(self):
        caminho_export = os.path.join(DIRETORIO_RELATORIO_BAIXADO, "remessa_baixada.csv")
        df_relatorio = pd.read_csv(caminho_export, sep=";", encoding="utf-8")

        df_despesa = df_relatorio[df_relatorio["Natureza"] == "D"].copy()

        for col in ("Vlr. Bruto",):
            if col in df_despesa.columns:
                df_despesa[col] = (
                    df_despesa[col].astype(str).str.strip()
                    .str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
                )
                df_despesa[col] = pd.to_numeric(df_despesa[col], errors="coerce")

        # Soma por operadora (cobre o criterio "repeticao para todas as operadoras")
        df_soma_operadora = df_despesa.groupby("Grp. Oper .Prest")["Vlr. Bruto"].sum().reset_index()
        df_soma_operadora.rename(columns={"Vlr. Bruto": "Vlr_Bruto_AGI"}, inplace=True)

        # Para cada operadora, busca o "Subtotal Despesa" na ABA correspondente do Encontro
        # de Contas (confirmado por print: 1 aba por operadora, ex. "AMPERNET").
        df_soma_operadora["Grp. Oper .Prest"] = df_soma_operadora["Grp. Oper .Prest"].astype(str).str.upper().str.strip()
        # to_numeric com errors="coerce": quando _ler_subtotal_despesa_ec retorna None (arquivo
        # do EC ausente, aba nao encontrada, celula vazia, etc.) vira NaN em vez de None puro -
        # None deixava a coluna com dtype "object" e o .abs() mais abaixo quebrava com
        # "TypeError: bad operand type for abs(): 'NoneType'".
        df_soma_operadora["Vlr_Subtotal_EC"] = pd.to_numeric(
            df_soma_operadora["Grp. Oper .Prest"].apply(self._ler_subtotal_despesa_ec), errors="coerce"
        )

        df_comparado = df_soma_operadora.copy()
        # Sinal: EC guarda despesa como valor NEGATIVO (ex.: -521,60); AGI soma como
        # positivo (521,6) - normaliza para o mesmo sinal antes de comparar (usa abs()).
        df_comparado["Diferenca"] = (df_comparado["Vlr_Bruto_AGI"].abs() - df_comparado["Vlr_Subtotal_EC"].abs()).round(2)
        # NaN em Diferenca = nao foi possivel comparar (EC nao encontrado p/ essa operadora) -
        # tratamos como inconsistente por seguranca, para nao passar batido sem alerta.
        df_comparado["Inconsistente"] = df_comparado["Diferenca"].isna() | (df_comparado["Diferenca"].abs() > 0.01)  # TODO: confirmar limiar de tolerancia oficial

        return df_comparado

    # ------------------------------------------------------------------
    # NOVO - le a celula "Subtotal Despesa" (CELULA_SUBTOTAL_DESPESA, confirmado = O87)
    # na aba do Encontro de Contas cujo nome bate com a operadora (ex.: "AMPERNET").
    # Confirmado por print real (imagem 10 do As Is): a aba tem um bloco "ENCONTRO DE
    # CONTAS" com linhas "Total Despesa" / "Total Contestação De..." / "Subtotal Despesa".
    # ------------------------------------------------------------------
    def _ler_subtotal_despesa_ec(self, operadora: str):
        if not DIRETORIO_ENCONTRO_CONTAS or not os.path.isfile(DIRETORIO_ENCONTRO_CONTAS):
            print("[EC] arquivo do Encontro de Contas nao configurado - ver DIRETORIO_ENCONTRO_CONTAS no config.py")
            return None
        try:
            wb = openpyxl.load_workbook(DIRETORIO_ENCONTRO_CONTAS, data_only=True)
        except Exception as e:
            print(f"[EC][ERRO] falha ao abrir '{DIRETORIO_ENCONTRO_CONTAS}': {e}")
            return None

        # Procura a aba cujo nome contem o nome da operadora (case-insensitive) - TODO:
        # confirmar se o nome da aba e EXATAMENTE igual ao "Grp. Oper .Prest" do AGI ou
        # se precisa de um DE-PARA (ex.: "AMPERNET" no AGI x "Ampernet Telecom" na aba).
        aba = None
        for nome in wb.sheetnames:
            if operadora.upper() in nome.upper():
                aba = nome
                break
        if aba is None:
            print(f"[EC][AVISO] nenhuma aba encontrada para operadora '{operadora}' em {wb.sheetnames}")
            return None

        try:
            valor = wb[aba][CELULA_SUBTOTAL_DESPESA].value
            return float(valor) if valor is not None else None
        except Exception as e:
            print(f"[EC][ERRO] falha ao ler célula {CELULA_SUBTOTAL_DESPESA} da aba '{aba}': {e}")
            return None

    # ------------------------------------------------------------------
    # NOVO - nao existe equivalente no exemplo (la so gravava um Excel local, sem alerta).
    # TODO: decidir o mecanismo real de sinalizacao (Webfat / e-mail / log em banco) e
    # implementar aqui. Por ora, so grava um Excel de divergencias e loga no console.
    # ------------------------------------------------------------------
    def _sinalizar_inconsistencias(self, df_comparado):
        divergentes = df_comparado[df_comparado["Inconsistente"] == True]
        if divergentes.empty:
            print("[HU-20] Nenhuma divergencia encontrada entre AGI e Encontro de Contas.")
            return

        os.makedirs(DIRETORIO_INCONSISTENCIAS, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho = os.path.join(DIRETORIO_INCONSISTENCIAS, f"inconsistencias_{PERIODO_REF}_{timestamp}.xlsx")
        divergentes.to_excel(caminho, index=False)
        print(f"[HU-20][INCONSISTENCIA] {len(divergentes)} operadora(s) divergente(s) -> {caminho}")

        # TODO [NOVO]: alem de gravar o Excel, sinalizar de fato (ex.: gravar em
        # tbl_rpa_log_detraf_despesa_arquivos com status='INCONSISTENTE', ou enviar e-mail,
        # ou atualizar status no Webfat). Nenhuma dessas integracoes existe hoje no projeto.


if "__main__" == __name__:
    Verificacao_Relatorio().Fluxo_Verificacao_Relatorio()
